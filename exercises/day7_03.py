#tabliczka
import openpyxl as xl
# excel = xl.Workbook('tabliczka_mnozenia.xlsx')
# excel.save('tabliczka_mnozenia.xlsx')
size = int(input('Podaj rozmiar tabeli'))
def table(size):
    excel = xl.load_workbook('tabliczka_mnozenia.xlsx')
    arkusz = excel.active
    for row in arkusz:
        for cell in row:
            cell.value = None
    for i in range(1, size + 1):
        for y in range(1, size + 1):
            komorka = arkusz.cell(i, y)
            komorka.value = i * y
    excel.save('tabliczka_mnozenia.xlsx')
table(size)