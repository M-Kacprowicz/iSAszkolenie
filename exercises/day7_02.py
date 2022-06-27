#excele
import openpyxl as xl
# excel = xl.Workbook()
# excel.save('arkusz_testowy.xlsx')
excel = xl.load_workbook('arkusz_testowy.xlsx')
print(excel.sheetnames)
arkusz = excel.active
komorkaA1 = arkusz.cell(1, 1)
print(komorkaA1.column_letter)
print(komorkaA1.value)
komorkaA2 = arkusz.cell(2, 1)
komorkaA2.value = 'komorka A2'
print(komorkaA2.value)
excel.save('arkusz_testowy.xlsx')
# arkusz = excel.active
# arkusze = excel.sheetnames
# arkusz = excel['Arkusz1']
# print(komorka.value)
# komorka.value = 'Nowa wartość'
# excel.save('plik.xlsx')